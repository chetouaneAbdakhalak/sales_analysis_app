import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import tkinter as tk
from tkinter import filedialog, messagebox

# ---------- 1. LOAD DATA ----------
def load_data(path):
    """Load sales data from CSV, Excel, or Parquet"""
    if path.endswith(".csv"):
        return pd.read_csv(path)
    elif path.endswith(".xlsx"):
        return pd.read_excel(path)
    elif path.endswith(".parquet"):
        return pd.read_parquet(path)
    else:
        raise ValueError("Unsupported file format. Please use CSV, Excel, or Parquet.")
# ---------- 2. CLEAN DATA ----------
def clean_data(df):
    """Remove duplicates and fill missing values"""
    df2 = df.copy()
    df2 = df2.dropna(how="all")
    num_cols = df2.select_dtypes(include=["number"]).columns
    df2[num_cols] = df2[num_cols].fillna(0)
    obj_cols = df2.select_dtypes(include=["object"]).columns
    df2[obj_cols] = df2[obj_cols].fillna("Unknown")
    df2['Date'] = pd.to_datetime(df2["Date"])
    return df2

# ---------- 3. SUMMARIES ----------
def get_summaries(df):
    """Return product, category summaries and totals"""
    df["YearMonth"] = df["Date"].dt.to_period("M")
    product_summary = df.groupby("Product")["Total Sales"].sum().reset_index()
    category_summary = df.groupby("Category")["Total Sales"].sum().reset_index()
    top_categories = category_summary.sort_values(by="Total Sales", ascending=False).head(5)
    top_products = product_summary.sort_values(by="Total Sales", ascending=False).head(5)
    monthly_sales = df.groupby("YearMonth")["Total Sales"].sum().reset_index()
    totals = pd.DataFrame({
        "Total Sales": [round(df["Total Sales"].sum(),2)],
        "Average Sales": [round(df["Total Sales"].mean(),2)]
    })
    
    category_monthly = df.groupby(["YearMonth", "Category"])["Total Sales"].sum().reset_index()
    pivot_sales = category_monthly.pivot_table(
        index="YearMonth", 
        columns="Category", 
        values="Total Sales", 
        aggfunc="sum"
    )

    return product_summary, category_summary, top_categories, top_products, totals , monthly_sales, pivot_sales

# ---------- 4. PLOT CHART ----------
def plot_chart(product_summary,monthly_sales, chart_path_1, chart_path_2):
    """Create and save bar chart of product sales"""
    plt.figure(figsize=(8,4))
    plt.bar(product_summary["Product"], product_summary["Total Sales"], color="blue")
    plt.title("Sales by Product")
    plt.xlabel("Product")
    plt.ylabel("Total Sales")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(chart_path_1, dpi=100)  # save chart as image
    plt.close()
    plt.figure(figsize=(8,4))
    plt.plot(monthly_sales["YearMonth"].astype(str).tolist(),monthly_sales["Total Sales"].tolist(), color="blue")
    plt.title("Revenue by Month")
    plt.xlabel("Year Months")
    plt.ylabel("Total Sales")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(chart_path_2, dpi=100)  # save chart as image
    plt.close()  

# ---------- 5. SAVE TO EXCEL ----------
def save_to_excel(df, product_summary, category_summary,top_categories, top_products, totals, excel_path, chart_path_1 , chart_path_2, pivot_sales):
    """Save raw + summaries + totals + chart into Excel"""
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        product_summary.to_excel(writer, sheet_name="Product Summary", index=False)
        category_summary.to_excel(writer, sheet_name="Category Summary", index=False)
        top_categories.to_excel(writer, sheet_name="Top Categories", index=False)
        top_products.to_excel(writer, sheet_name="Top Products", index=False)
        totals.to_excel(writer, sheet_name="Grand Total", index=False)
        pivot_sales.to_excel(writer, sheet_name="Pivot Table", index=True)

    # Reopen Excel file and insert chart
    wb = load_workbook(excel_path)
    ws = wb.create_sheet("Charts")
    img = XLImage(chart_path_1)
    img.width, img.height = 600, 300  # resize chart
    ws.add_image(img, "B2")
    wb.save(excel_path)
    img = XLImage(chart_path_2)
    img.width, img.height = 600, 300  # resize chart
    ws.add_image(img, "B18")
    wb.save(excel_path)

# ---------- 6. MAIN ----------
# def main():
#     # Paths
#     csv_path = /
#     excel_path = /
#     chart_path_1 = /
#     chart_path_2 = /
#     # Run steps
#     df = load_data(csv_path)
#     df = clean_data(df)
#     product_summary, category_summary,top_categories, top_products, totals , monthly_sales , pivot_sales= get_summaries(df)
#     plot_chart(product_summary, monthly_sales,chart_path_1, chart_path_2)
#     save_to_excel(df, product_summary, category_summary,top_categories, top_products, totals, excel_path, chart_path_1 ,chart_path_2 ,  pivot_sales)

# if __name__ == "__main__":
#     main()



def process_file():
    csv_path = filedialog.askopenfilename(
        title="Select Sales Data File",
        filetypes=[
    ("CSV files", "*.csv"),
    ("Excel files", "*.xlsx"),
    ("Parquet files", "*.parquet")
]

    )
    if not csv_path:
        return  # user canceled

    # Generate output paths
    base = csv_path.rsplit(".", 1)[0]
    excel_path = base + "_report.xlsx"
    chart_path_1 = base + "_chart1.png"
    chart_path_2 = base + "_chart2.png"

    try:
        # Run pipeline
        df = load_data(csv_path)
        df = clean_data(df)
        product_summary, category_summary, top_categories, top_products, totals, monthly_sales, pivot_sales = get_summaries(df)
        plot_chart(product_summary, monthly_sales, chart_path_1, chart_path_2)
        save_to_excel(df, product_summary, category_summary, top_categories, top_products, totals, excel_path, chart_path_1, chart_path_2, pivot_sales)

        messagebox.showinfo("Success", f"Report saved:\n{excel_path}")
        root.destroy()  # close app
    except Exception as e:
        messagebox.showerror("Error", str(e))
        root.destroy()

# ---- Welcome Page ----
root = tk.Tk()
root.title("Sales Report Generator")
root.geometry("450x300")

welcome_label = tk.Label(root, text="Welcome!\n\nClick below to choose your sales file.\n Only csv , excel and parquet files are supported", font=("Arial", 14), pady=20)
welcome_label.pack()

choose_btn = tk.Button(root, text="Choose File", font=("Arial", 11, "bold"), command=process_file, width=15, height=2, bg="lightblue")
choose_btn.pack(pady=20)

root.mainloop()
